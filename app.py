from dotenv import load_dotenv
import os
from tkinter import *
from tkcalendar import Calendar, DateEntry
import pandas as pd
from datetime import date, timedelta
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles.borders import Border, Side
from sqlalchemy import create_engine, text
from PIL import ImageTk, Image
import matplotlib.pyplot as plt
from matplotlib.sankey import Sankey


load_dotenv()

#define tk window
windowMain = Tk()
windowMain.title('BudgetIO')
windowMain.geometry('600x400')

#DATE INPUT
dateText = Label(windowMain, text='Enter date of purchase:')
dateText.grid(row=1, column=1)
Bdate = DateEntry(windowMain,date_pattern='yyyy-mm-dd', foreground='white', background='white', width=19)
Bdate.grid(column=2, row=1)

#PRICE INPUT
priceText = Label(windowMain, text="Enter price of item/service:")
priceText.grid(row=2, column=1)
priceVar = StringVar()
price = Entry(windowMain, textvariable=priceVar, background='white', foreground='black')
price.grid(row=2,column=2)

#PLACE OF PURCHASE INPUT
whereText = Label(windowMain, text="Enter place of purchase:")
whereText.grid(row=3, column=1)
whereVar = StringVar()
where = Entry(windowMain, textvariable=whereVar, background='white', foreground='black')
where.grid(row=3, column=2)

#CATEGORY OF PURCHASE INPUT
categoriesText = Label(windowMain, text='Enter category of purchase:')
categoriesText.grid(row=4, column=1)
categories = [
    'Housing',
    'Utilities',
    'Grocery',
    'Phone',
    'Fun',
    'Misc'
    ]
categoryVar = StringVar()
categoryDrop = OptionMenu(windowMain, categoryVar, *categories)
categoryDrop.config(width=19)
categoryDrop.grid(row=4, column=2)

#SUCCESS LABEL TO SHOW AFTER INSERTION OF EXPENSE
successLabelVar = StringVar()
successLabel = Label(windowMain, text='')
successLabel.grid(row=5, column=3)

#INSERT DATA BUTTON
insertGo = Button(windowMain,text='Insert', command=(lambda: (insertExpenseSQL(Bdate, priceVar.get(), whereVar.get(), categoryVar.get(), successLabel))))
insertGo.grid(row=4, column=3)

logoImage = Image.open("/Users/alexbrady/Library/Mobile Documents/com~apple~CloudDocs/Budget Repo/Budget/Images/BudgetIOLogo.png")
logoImage = logoImage.resize((300,250))
convertLogoImage = ImageTk.PhotoImage(logoImage)
logoImageLabel = Label(windowMain, image=convertLogoImage, justify='left')
logoImageLabel.grid(row=5, column=1)


#INSERT INCOME WINDOW
def newInsertIncomeWindow():
    incomeWindow = Toplevel(windowMain)
    incomeWindow.title('Income')
    incomeWindow.geometry('600x400')

    #DATE OF INCOME
    incomeDateLabel = Label(incomeWindow, text='Enter date of income:')
    incomeDateLabel.grid(row=1, column=1)

    incomeDateEntry = DateEntry(incomeWindow, date_pattern='yyyy-mm-dd', width=19)
    incomeDateEntry.grid(row=1, column=2)

    #CHECKING AMOUNT
    checkingLabel = Label(incomeWindow, text='Enter checking amount:')
    checkingLabel.grid(row=2, column=1)

    checkingVar = StringVar()
    checkingEntry = Entry(incomeWindow, textvariable=checkingVar, background='white', foreground='black')
    checkingEntry.grid(row=2, column=2)

    #SAVINGS AMOUNT
    savingsLabel = Label(incomeWindow, text='Enter savings amount:')
    savingsLabel.grid(row=3, column=1)

    savingsVar = StringVar()
    savingsEntry = Entry(incomeWindow, textvariable=savingsVar, background='white', foreground='black')
    savingsEntry.grid(row=3,column=2)

    #RETIREMENT AMOUNT
    retirementLabel = Label(incomeWindow, text='Enter retirement amount:')
    retirementLabel.grid(row=4, column=1)

    retirementVar = StringVar()
    retirementEntry = Entry(incomeWindow, textvariable=retirementVar, background='white', foreground='black')
    retirementEntry.grid(row=4, column=2)

    #TOTAL COMP AMOUNT
    totalCompLabel = Label(incomeWindow, text='Enter total comp pre-tax:', background='white', foreground='black')
    totalCompLabel.grid(row=5, column=1)

    totalCompVar = StringVar()
    totalComp = Entry(incomeWindow, textvariable=totalCompVar)
    totalComp.grid(row=5, column=2)

    #TAX AMOUNT
    taxLabel = Label(incomeWindow, text='Enter tax amount:', background='white', foreground='black')
    taxLabel.grid(row=6, column=1)

    taxVar = StringVar()
    tax = Entry(incomeWindow, textvariable=taxVar)
    tax.grid(row=6, column=2)


    #SOURCE OF INCOME
    sourceLabel = Label(incomeWindow, text='Enter source of income:')
    sourceLabel.grid(row=7, column=1)

    sources = [
    'Qorvo',
    'Other'
    ]
    sourceVar = StringVar()
    sourceDrop = OptionMenu(incomeWindow, sourceVar, *sources)
    sourceDrop.config(width=19)
    sourceDrop.grid(row=7, column=2)

    successLabelVar = StringVar()
    successLabel = Label(incomeWindow, text='')
    successLabel.grid(row=7, column=3)


    insertIncomeButton = Button(incomeWindow, text='Insert', command= lambda: insertIncomeSQL(incomeDateEntry, checkingVar.get(), savingsVar.get(), retirementVar.get(), sourceVar.get(), successLabel, totalCompVar.get(), taxVar.get()))
    insertIncomeButton.grid(row=8,column=3)

#VIEW-RANGE WINDOW
def newExpenseViewWindow():
    rangeWindow = Toplevel(windowMain)
    rangeWindow.title('Expense Export')
    rangeWindow.geometry('600x400')

    #START DATE OF SEARCH
    lDateLable = Label(rangeWindow, text="DATE: From:")
    lDate = DateEntry(rangeWindow,date_pattern='yyyy-mm-dd', foreground='white', background='white', width=19)
    lDateLable.grid(column=1, row=1)
    lDate.grid(column=2, row=1)

    #END DATE OF SEARCH
    rDateLabel = Label(rangeWindow, text="To:")
    rDate = DateEntry(rangeWindow,date_pattern='yyyy-mm-dd', foreground='white', background='white', width=19)
    rDateLabel.grid(column=3, row=1)
    rDate.grid(column=4, row=1)

    #START PRICE INPUT
    lPriceVar = StringVar()
    lPriceLable = Label(rangeWindow, text='PRICE: From:')
    lPrice = Entry(rangeWindow, textvariable=lPriceVar, background='white', foreground='black')
    lPriceLable.grid(column=1, row=2)
    lPrice.grid(column=2,row=2)

    #END PRICE INPUT
    rPriceVar = StringVar()
    rPriceLabel = Label(rangeWindow, text='To:')
    rPrice = Entry(rangeWindow, textvariable=rPriceVar, background='white', foreground='black')
    rPriceLabel.grid(column=3, row=2)
    rPrice.grid(column=4, row=2)

    #CATEGORY INPUT
    categoriesText = Label(rangeWindow, text='Enter category of purchase:')
    categoriesText.grid(row=3, column=1)
    categories = [
        'Housing',
        'Utilities',
        'Grocery',
        'Phone',
        'Fun',
        'Misc',
        'Home'
        ]
    categoryVarRange = StringVar()
    categoryDrop = OptionMenu(rangeWindow, categoryVarRange, *categories)
    categoryDrop.config(width=19)
    categoryDrop.grid(row=3, column=2)

    #EXECUTE RANGE SELECTION BUTTON
    executeRange = Button(rangeWindow, text='Execute', command= lambda:viewExpenseRangeSQL(lDate, rDate, lPriceVar.get(), rPriceVar.get(), categoryVarRange.get(), csvVar.get()))
    executeRange.grid(row=4, column=4)

    executeWideOpen = Button(rangeWindow, text='Wide Open', command=lambda:expenseSQLWideOpen(csvVar.get()))
    executeWideOpen.grid(row=5,column=4)

    #CSV NAME
    csvVar = StringVar()
    csvLabel = Label(rangeWindow, text='*File name:')
    csvName = Entry(rangeWindow, textvariable=csvVar, background='white', foreground='black')
    csvLabel.grid(column=1, row=4)
    csvName.grid(column=2, row=4)

def newViewIncomeWindow():
    viewIncomeWindow = Toplevel(windowMain)
    viewIncomeWindow.title('Income Export')
    viewIncomeWindow.geometry('300x100')

    csvLabel = Label(viewIncomeWindow, text='*File name:')
    csvLabel.grid(row=1, column=1)

    csvVar = StringVar()
    csvEntry = Entry(viewIncomeWindow, textvariable=csvVar, background='white', foreground='black')
    csvEntry.grid(row=1, column=2)

    executeIncomeExportButton = Button(viewIncomeWindow, text='Execute', command= lambda: viewSQLIncome(csvVar.get()))
    executeIncomeExportButton.grid(row=3,column=2)

def newDeleteIncomeWindow():
    deleteIncomeWindow = Toplevel(windowMain)
    deleteIncomeWindow.title('Delete Income Entry')
    deleteIncomeWindow.geometry('600x400')

    #DATE TO DELETE
    dateText = Label(deleteIncomeWindow, text='Enter date of income')
    dateText.grid(row=1, column=1)
    Bdate = DateEntry(deleteIncomeWindow,date_pattern='yyyy-mm-dd', foreground='white', background='white', width=19)
    Bdate.grid(column=2, row=1)

    #CHECKING AMOUNT TO DELETE
    checkingLabel = Label(deleteIncomeWindow, text="Enter checking amount:")
    checkingLabel.grid(row=2, column=1)
    checkingVar = StringVar()
    checking = Entry(deleteIncomeWindow, textvariable=checkingVar, background='white', foreground='black')
    checking.grid(row=2,column=2)

    #SAVINGS AMOUNT TO DELETE
    savingsLabel = Label(deleteIncomeWindow, text="Enter savings amount:")
    savingsLabel.grid(row=3, column=1)
    savingsVar = StringVar()
    savings = Entry(deleteIncomeWindow, textvariable=savingsVar, background='white', foreground='black')
    savings.grid(row=3,column=2)

    #RETIREMENT AMOUNT TO DELETE
    retirementLabel = Label(deleteIncomeWindow, text="Enter retirement amount:")
    retirementLabel.grid(row=4, column=1)
    retirementVar = StringVar()
    retirement = Entry(deleteIncomeWindow, textvariable=retirementVar, background='white', foreground='black')
    retirement.grid(row=4,column=2)

    #SOURCE TO DELETE
    sourceLabel = Label(deleteIncomeWindow, text='Enter source of income:')
    sourceLabel.grid(row=5, column=1)

    sources = [
    'Qorvo',
    'Other'
    ]
    sourceVar = StringVar()
    sourceDrop = OptionMenu(deleteIncomeWindow, sourceVar, *sources)
    sourceDrop.config(width=19)
    sourceDrop.grid(row=5, column=2)

    #DELETE INCOME BUTTON
    deleteIncomeButton = Button(deleteIncomeWindow, text='Delete', command=lambda: newVerifyDeleteIncomeWindow(deleteIncomeWindow, Bdate, checkingVar, savingsVar, retirementVar, sourceVar))
    deleteIncomeButton.grid(row=5, column=3)

def newVerifyDeleteIncomeWindow(window, date : Calendar, checking, savings, retirement, source):
    verifyDeleteIncomeWindow = Toplevel(window)
    verifyDeleteIncomeWindow.title('Verify Deletion')
    verifyDeleteIncomeWindow.geometry('350x150')

    total = int(checking.get()) + int(savings.get()) + int(retirement.get())
    verifyQuestion = Label(verifyDeleteIncomeWindow,text=f"Do you wish to delete income entry of {total} reported on {date.get_date()}?")
    verifyQuestion.grid(row=1,column=1)

    yesButton = Button(verifyDeleteIncomeWindow,text='YES', command=lambda: sqlIncomeDelete(date, checking, savings, retirement, source))
    yesButton.grid(row=2, column=1)

    noButton = Button(verifyDeleteIncomeWindow,text='NO', command= lambda : verifyDeleteIncomeWindow.destroy(), justify='left')
    noButton.grid(row=2, column=2)  

def newDeleteExpenseWindow():
    deleteExpenseWindow = Toplevel(windowMain)
    deleteExpenseWindow.title('Delete Expense Entry')
    deleteExpenseWindow.geometry('600x400')

    #DATE TO DELETE
    dateText = Label(deleteExpenseWindow, text='Enter date of purchase:')
    dateText.grid(row=1, column=1)
    Bdate = DateEntry(deleteExpenseWindow,date_pattern='yyyy-mm-dd', foreground='white', background='white', width=19)
    Bdate.grid(column=2, row=1)

    #PRICE TO DELETE
    priceText = Label(deleteExpenseWindow, text="Enter price of item/service:")
    priceText.grid(row=2, column=1)
    priceVar = StringVar()
    price = Entry(deleteExpenseWindow, textvariable=priceVar, background='white', foreground='black')
    price.grid(row=2,column=2)

    #PLACE OF PURCHASE TO DELETE
    whereText = Label(deleteExpenseWindow, text="Enter place of purchase:")
    whereText.grid(row=3, column=1)
    whereVar = StringVar()
    where = Entry(deleteExpenseWindow, textvariable=whereVar, background='white', foreground='black')
    where.grid(row=3, column=2)

    #CATEGORY OF PURCHASE TO DELETE
    categoriesText = Label(deleteExpenseWindow, text='Enter category of purchase:')
    categoriesText.grid(row=4, column=1)
    categories = [
        'Housing',
        'Utilities',
        'Grocery',
        'Phone',
        'Fun',
        'Misc'
        ]
    categoryVar = StringVar()
    categoryDrop = OptionMenu(deleteExpenseWindow, categoryVar, *categories)
    categoryDrop.config(width=19)
    categoryDrop.grid(row=4, column=2)

    #DELETE DATA BUTTON
    deleteGo = Button(deleteExpenseWindow,text='Delete', command = lambda: newVerifyDeleteExpenseWindow(deleteExpenseWindow, Bdate.get_date(), priceVar, whereVar, categoryVar))
    deleteGo.grid(row=6, column=3)

def newVerifyDeleteExpenseWindow(topWindow, when : Calendar, price : StringVar, where : StringVar, category : StringVar):
    verifyDeleteExpenseWindow = Toplevel(topWindow)
    verifyDeleteExpenseWindow.title('Verify Deletion')
    verifyDeleteExpenseWindow.geometry('350x150')

    verifyQuestion = Label(verifyDeleteExpenseWindow,text=f'Do you wish to delete this entry of ${price.get()} spent at {where.get()} on {when}?', justify='center', wraplength=260)
    verifyQuestion.grid(row=1,column=1)

    yesButton = Button(verifyDeleteExpenseWindow,text='YES', command=lambda: (sqlExpenseDelete(when, price.get(), where.get(), category.get())), justify='left')
    yesButton.grid(row=2, column=1)

    noButton = Button(verifyDeleteExpenseWindow,text='NO', command= lambda : verifyDeleteExpenseWindow.destroy(), justify='left')
    noButton.grid(row=2, column=2)


#INCOME EXPENSE REPORT WINDOW
def newExpenseToIncomeWindow():
    expenseToIncomeWindow = Toplevel(windowMain)   
    expenseToIncomeWindow.title('Expense To Income Report')
    expenseToIncomeWindow.geometry('600x400')

    #FROM   
    lDateLable = Label(expenseToIncomeWindow, text='From:')
    lDateLable.grid(row=1, column=1)

    lDate = DateEntry(expenseToIncomeWindow, date_pattern='yyyy-mm-dd')
    lDate.grid(row=1, column=2)

    #TO
    rDateLabel = Label(expenseToIncomeWindow, text='To:')
    rDateLabel.grid(row=1, column=3)

    rDate = DateEntry(expenseToIncomeWindow, date_pattern='yyyy-mm-dd')
    rDate.grid(row=1, column=4)

    #SHOW EXPENSE TOTAL LABEL
    expenseLabel = Label(expenseToIncomeWindow, text='Total expenses:')
    expenseLabel.grid(row=3, column=1)
    expenseTotal = Label(expenseToIncomeWindow, text='')
    expenseTotal.grid(row=3, column=2)

    #SHOW INCOME TOTAL LABLE
    incomeLabel = Label(expenseToIncomeWindow, text='Total Income:')
    incomeLabel.grid(row=4, column=1)
    incomeTotal = Label(expenseToIncomeWindow, text='')
    incomeTotal.grid(row=4, column=2)

    #SHOW INCOME MINUS EXPENSE LABEL
    netLabel = Label(expenseToIncomeWindow, text='Net:')
    netLabel.grid(row=5, column=1)
    net = Label(expenseToIncomeWindow, text='')
    net.grid(row=5, column=2)

    #TODO CONFIGURE CHECKBOX EXCEL OPTION FOR REPORTS
    #EXCEL CHECKBOX OPTION
    createExcelVar = IntVar()
    createExcelCheckBox = Checkbutton(expenseToIncomeWindow, variable=createExcelVar, text='Create Excel')
    createExcelCheckBox.grid(row=2, column=5)

    #EXECUTE EXPENSE TO INCOME REPORT
    executeExpenseToIncomeReportButton = Button(expenseToIncomeWindow, text='Execute', command= lambda: viewExpenseToIncomeReport(lDate, rDate, incomeTotal, expenseTotal, net)) #PASS TOTAL LABELS IN FUNC FOR UPDATE
    executeExpenseToIncomeReportButton.grid(row=3, column=5)

def newSankeyReportWindow():
    sankeyWindow = Toplevel(windowMain)
    sankeyWindow.title('Sankey Diagram')
    sankeyWindow.geometry('600x400')

    lDateLabel = Label(sankeyWindow, text='From:')
    lDateLabel.grid(row=1, column=1)
    lDate = DateEntry(sankeyWindow, date_pattern='yyyy-mm-dd')
    lDate.grid(row=1, column=2)

    rDateLabel = Label(sankeyWindow, text='To:')
    rDateLabel.grid(row=1, column=3)
    rDate = DateEntry(sankeyWindow, date_pattern='yyyy-mm-dd')
    rDate.grid(row=1, column=4)

    sankeyChartExecuteButton = Button(sankeyWindow, text='Execute', command=lambda: SankeyChart(lDate, rDate))
    sankeyChartExecuteButton.grid(row=2, column=4)


#MENU BAR TKINTER
mainMenu = Menu(windowMain)
exportMenu = Menu(mainMenu)
deleteMenu = Menu(mainMenu)
insertMenu = Menu(mainMenu)
reportMenu = Menu(mainMenu)

mainMenu.add_cascade(label='Export', menu=exportMenu)
exportMenu.add_command(label='Expense', command=newExpenseViewWindow)
exportMenu.add_separator()
exportMenu.add_command(label='Income', command=newViewIncomeWindow)
mainMenu.add_cascade(label='Delete', menu=deleteMenu)
deleteMenu.add_separator()
deleteMenu.add_command(label='Expense', command=newDeleteExpenseWindow)
deleteMenu.add_separator()
deleteMenu.add_command(label='Income', command=newDeleteIncomeWindow)
mainMenu.add_cascade(label='Insert', menu=insertMenu)
insertMenu.add_command(label='Income', command=newInsertIncomeWindow)
mainMenu.add_cascade(label='Report', menu=reportMenu)
reportMenu.add_command(label='Expense to Income', command=newExpenseToIncomeWindow)
reportMenu.add_command(label='Weekly Spending Report', command= lambda: weeklySpendingReport())
reportMenu.add_command(label='Sankey', command=newSankeyReportWindow)

windowMain.config(menu=mainMenu)


#--------SQL METHODS----------#

#INSERTS PURCHASES INTO DB -------METHOD--------
def insertExpenseSQL(bdate : Calendar, amount , Where , Category, success : Label):
    try:
        cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
        with cxn.connect() as con:
            con.execute(text((f"INSERT INTO budget.expenses VALUES ('{bdate.get_date()}', {amount}, '{Where}', '{Category}')")))
            con.commit()
        success.config(text='Success!', foreground='green')
        success.update()
    except:
        success.config(text='Error!', foreground='red')
        success.update()


#INSERT INCOME TO SQL
def insertIncomeSQL(date : Calendar, checking, savings, retirement, source, success : Label, totalComp, tax):
    try:
        cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
        with cxn.connect() as con:
            con.execute(text(f"INSERT INTO budget.income VALUES ('{date.get_date()}', {int(checking)}, {int(savings)}, {int(retirement)}, '{source}', {int(totalComp)}, {int(tax)})"))
            con.commit()
        success.config(text='Success!', foreground='green')
        success.update()
    except:
        success.config(text='Error!', foreground='red')
        success.update()
    
def viewSQLIncome(xlName):
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")   
    with cxn.connect() as con:
            query = "SELECT * FROM Budget.income;"
            df = pd.read_sql(query, cxn)
        
    df.to_excel(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {date.today()}.xlsx")
    workbook = openpyxl.load_workbook(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {date.today()}.xlsx")
    sheet = workbook.active
    #CONFIGURE DATE COLUMN IN EXCEL
    dateCol = sheet.column_dimensions['B']
    dateCol.number_format = 'YYYY MM DD'

    workbook.save(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {date.today()}.xlsx")


#SEARCHES SQL EXPENSE TABLE VIA RANGE -----METHOD--------
def viewExpenseRangeSQL(ldate : Calendar, rdate : Calendar, lprice, rprice, category, xlName):
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
   

    #RANGE BY DATE, CATEGORY, AND PRICE
    if (ldate != '' and rdate != '') and lprice != '' and rprice !='' and category != '':
        query = f"SELECT Bdate, price, location, category FROM budget.expenses WHERE bdate BETWEEN '{ldate.get_date()}' AND '{rdate.get_date()}' AND category = '{category}' AND price BETWEEN {lprice} AND {rprice}"
        df = pd.read_sql(query, cxn)
        
    #RANGE BY DATE AND CATEGORY
    if (ldate != '' and rdate != '') and lprice == '' and rprice == '' and category != '':
        query = f"SELECT Bdate, price, location, category FROM budget.expenses WHERE bdate BETWEEN '{ldate.get_date()}' AND '{rdate.get_date()}' AND category = '{category}';"
        df = pd.read_sql(query, cxn)

    #RANGE BY DATE AND PRICE
    if (ldate != '' and rdate != '') and lprice != '' and rprice != '' and category == '':
        query = f"SELECT Bdate, price, location, category FROM budget.expenses WHERE bdate BETWEEN '{ldate.get_date()}' AND '{rdate.get_date()}' AND price BETWEEN {lprice} AND {rprice};"
        df = pd.read_sql(query, cxn)

    #RANGE ONLY BY DATE
    if (ldate != '' and rdate != '') and lprice == '' and rprice == '' and category == '':
        query = f"SELECT Bdate, price, location, category FROM budget.expenses WHERE bdate BETWEEN '{ldate.get_date()}' AND '{rdate.get_date()}';"
        df = pd.read_sql(query, cxn)
    
        
    df.to_excel(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {ldate.get_date()} {rdate.get_date()}.xlsx")

    #CREATING PIE CHART IN EXCEL
    workbook = openpyxl.load_workbook(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {ldate.get_date()} {rdate.get_date()}.xlsx")
    sheet = workbook.active
    chart = PieChart()
    labels = Reference(sheet, min_col=5, max_col=5, min_row=2, max_row=len(df.index)+1)
    data = Reference(sheet, min_col=3, max_col=3, min_row=2, max_row=len(df.index)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = 'Categorical Spending'
    sheet.add_chart(chart, 'G2')

    #FORMATTING DATE COLUMN
    dateCol = sheet.column_dimensions['B']
    dateCol.number_format = 'YYYY MM DD'

    #ADDING SUMMATION CELL TO PRICE COLUMN
    sheet[f'C{len(df.index)+2}'] = f'=SUM(C2:C{len(df.index)+1})'
    sheet[f'A{len(df.index)+2}'] = 'SUM'
    #BORDER FOR SUMMATION CELL
    thickBorder = Border(top = Side(style = 'thick'))
    sheet.cell(row = len(df.index)+2, column = 3).border = thickBorder

    workbook.save(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {ldate.get_date()} {rdate.get_date()}.xlsx")

#FOR RANGE PAGE - WIDE OPEN ----- METHOD ----------
def expenseSQLWideOpen(xlName):
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
    
    query = "SELECT * FROM Budget.expenses;"
    df = pd.read_sql(query, cxn)
    
        
    df.to_excel(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {date.today()}.xlsx")
    
    
    #CREATING PIE CHART IN EXCEL
    workbook = openpyxl.load_workbook(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {date.today()}.xlsx")
    sheet = workbook.active
    chart = PieChart()
    labels = Reference(sheet, min_col=5, max_col=5, min_row=2, max_row=len(df.index)+1)
    data = Reference(sheet, min_col=3, max_col=3, min_row=2, max_row=len(df.index)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = 'Categorical Spending'
    sheet.add_chart(chart, 'G2')
    
    
    #FORMATTING DATE COLUMN
    dateCol = sheet.column_dimensions['B']
    dateCol.number_format = 'YYYY MM DD'

    #ADDING SUMMATION CELL TO PRICE COLUMN
    sheet[f'C{len(df.index)+2}'] = f'=SUM(C2:C{len(df.index)+1})'
    sheet[f'A{len(df.index)+2}'] = 'SUM'
    #BORDER FOR SUMMATION CELL
    thickBorder = Border(top = Side(style = 'thick'))
    sheet.cell(row = len(df.index)+2, column = 3).border = thickBorder

    workbook.save(f"{os.getenv('BUDGETIO_OUTPUT_PATH')}{xlName} {date.today()}.xlsx")
    

def sqlExpenseDelete(bdate : Calendar, amount , Where , Category):
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
    with cxn.connect() as con:
        con.execute(text(f"DELETE FROM expenses WHERE bdate = '{bdate}' AND price = {amount} AND location = '{Where}' AND category ='{Category}'"))
        con.commit()
    
        

def sqlIncomeDelete(date : Calendar, checking, savings, retirement, source):
    
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
    with cxn.connect() as con:
        con.execute(text(f"DELETE FROM income WHERE idate = '{date.get_date()}' AND checking = {checking.get()} AND savings = {savings.get()} AND retirement ={retirement.get()} AND source = '{source.get()}'"))
        con.commit() 
            

    


#INCOME EXPENSE SQL STATEMENT TO EXCEL FILE - ANALYSIS
def viewExpenseToIncomeReport(lDate : Calendar, rDate : Calendar, incomeLabel : Label, expenseLabel : Label, netLabel : Label):
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
    with cxn.connect() as con:
            
        incomeQuery = f"SELECT savings, checking FROM budget.income WHERE idate BETWEEN '{lDate.get_date()}' AND '{rDate.get_date()}'"
        incomeDF = pd.read_sql(incomeQuery, cxn)
        print(incomeDF)
        incomeTotal = sum(incomeDF['savings'])
        
        incomeLabel.config(text=f'{incomeTotal}')
        incomeLabel.update()

        expenseQuery = f"SELECT price FROM budget.expenses WHERE Bdate BETWEEN '{lDate.get_date()}' AND '{rDate.get_date()}'"
        expenseDF = pd.read_sql(expenseQuery, cxn)
        expenseTotal = expenseDF['price'].sum()
        expenseLabel.config(text=f'{expenseTotal}')
        expenseLabel.update()

        
        incomeMinusExpense = round(incomeTotal - expenseTotal, 2)

        if incomeMinusExpense < 0:
            netColor = 'red'
        if incomeMinusExpense >= 0:
            netColor = 'green'

        netLabel.config(text=f'{str(incomeMinusExpense)}', foreground=netColor)
        netLabel.update()

        
def weeklySpendingReport():
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
    with cxn.connect() as con:
        
        spendingQuery = f"SELECT category, price FROM budget.expenses WHERE Bdate BETWEEN '{date.today() - timedelta(days=7)}' AND '{date.today()}' ORDER BY Bdate"
        spendingDF = pd.read_sql(spendingQuery, cxn)
    
        #GRAPHING TO MATPLOTLIB WINDOW

        #DYNAMICALLY GETTING NUMBER OF CATEGOREIS FOR EXPLODE
        uniqueExpenseCategories = len(pd.unique(spendingDF['category']))
        explode = []
        i = 0
        while i < uniqueExpenseCategories:
            explode.append(0.1)
            i +=1

        plt.pie(spendingDF['price'], labels=spendingDF['category'], shadow = True, autopct = '%1.1f%%', explode=explode)
        plt.legend()
        plt.title(f'Spending Report {date.today() - timedelta(days=7)} To {date.today()}')
    
        plt.show()
        

def SankeyChart(lDate : Calendar, rDate : Calendar):
    cxn = create_engine(url=f"mysql+pymysql://root:{os.getenv('PASSWORD')}@localhost:3306/Budget")
    with cxn.connect() as con:
        incomeQuery = f"SELECT * FROM budget.income WHERE idate BETWEEN '{lDate.get_date()}' AND '{rDate.get_date()}';"
        incomeDF = pd.read_sql(incomeQuery, cxn)

        totalIncome = incomeDF['total_income'].sum()
        totalChecking = incomeDF['checking'].sum()
        totalSaving = incomeDF['savings'].sum()
        totalRetirement = incomeDF['retirement'].sum()
        totalTax = incomeDF['tax'].sum()
        TotalUnrealized = totalRetirement + totalTax
        totalRealized = totalIncome - TotalUnrealized

        expenseQuery = f"SELECT * FROM budget.expenses WHERE bdate BETWEEN '{lDate.get_date()}'AND '{rDate.get_date()}';"
        expenseDF = pd.read_sql(expenseQuery, cxn)
        
    sankeyLabels = [
    'Income',
    'Housing',
    'Utilities', #3
    'Grocery',
    'Phone',
    'Fun', #6
    'Misc',
    'Home'
    ]

    categories = [
    'Housing',
    'Utilities',
    'Grocery',
    'Phone',
    'Fun',
    'Misc',
    'Home'
    ]
    
    
    housingTotal = 0 
    utilitiesTotal = 0 
    groceryTotal = 0 
    phoneTotal = 0 
    funTotal = 0
    miscTotal = 0
    homeTotal = 0
    

    categoryList = [housingTotal, utilitiesTotal, groceryTotal, phoneTotal, funTotal, miscTotal, homeTotal]

    index = 0
    for category in categories: #for each category, query pandas by value and get sum of that column
        
        
        #filters dataframe by category value then summing price
        filteredDF = expenseDF[expenseDF['category'] == category]
        
        categoryList[index] = filteredDF['price'].sum()
        #categoryList[index] = filteredDF['price'].sum()  #sums price column
        
        index = index + 1

    Sankey(flows=[-totalIncome, categoryList[0], categoryList[1],
            categoryList[2], categoryList[3], categoryList[4],
            categoryList[5], categoryList[6], totalTax], labels=sankeyLabels, 
            orientations=[0, 1, -1, 0, 0, 0, 0, 0, 0], rotation=180).finish()
    
    plt.title(f'Sankey {lDate.get_date()} - {rDate.get_date()}')
    
    plt.show() 

    


windowMain.mainloop()